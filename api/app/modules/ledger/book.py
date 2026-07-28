"""The auxiliary book as a spreadsheet."""

from __future__ import annotations

import datetime as dt
from decimal import Decimal
from io import BytesIO
from typing import Final, Literal

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.ledger.schemas import AccountLedger

Locale = Literal["es", "en"]

#: The seven columns the ledger is specified to show, in that order. The
#: account is not one of them: it heads its own block of rows instead of being
#: repeated on every line.
HEADERS: Final[dict[Locale, list[str]]] = {
    "es": [
        "FECHA",
        "COMPROBANTE",
        "DESCRIPCIÓN",
        "TERCERO",
        "DÉBITO",
        "CRÉDITO",
        "SALDO",
    ],
    "en": [
        "DATE",
        "VOUCHER",
        "DESCRIPTION",
        "THIRD PARTY",
        "DEBIT",
        "CREDIT",
        "BALANCE",
    ],
}

LABELS: Final[dict[Locale, dict[str, str]]] = {
    "es": {
        "title": "Libro auxiliar",
        "total": "Totales cuenta",
        "grand_total": "Total general",
        "unbounded": "Sin límite",
        "empty": "Sin movimientos en el rango.",
    },
    "en": {
        "title": "Auxiliary ledger",
        "total": "Account total",
        "grand_total": "Grand total",
        "unbounded": "Unbounded",
        "empty": "No movements in the range.",
    },
}

WIDTHS: Final[list[int]] = [12, 13, 46, 34, 16, 16, 18]

MONEY_FORMAT: Final = "#,##0.00;[Red]-#,##0.00"
DATE_FORMAT: Final = "yyyy-mm-dd"

#: Title, who the books belong to, and what range was asked for. The headings
#: come next.
_HEADER_ROW: Final = 4

_BOLD: Final = Font(bold=True)


def build_auxiliary_book(
    accounts: list[AccountLedger],
    *,
    company: str,
    date_from: dt.date | None,
    date_to: dt.date | None,
    generated_at: dt.datetime,
    locale: Locale = "es",
) -> bytes:
    labels = LABELS[locale]

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = labels["title"]

    since = date_from.isoformat() if date_from else labels["unbounded"]
    until = date_to.isoformat() if date_to else labels["unbounded"]

    sheet.cell(row=1, column=1, value=labels["title"]).font = Font(bold=True, size=14)
    sheet.cell(row=2, column=1, value=company)
    sheet.cell(
        row=3,
        column=1,
        value=f"{since} — {until} · {generated_at:%Y-%m-%d %H:%M}",
    )

    for column, (heading, width) in enumerate(
        zip(HEADERS[locale], WIDTHS, strict=True), start=1
    ):
        sheet.cell(row=_HEADER_ROW, column=column, value=heading).font = _BOLD
        sheet.column_dimensions[get_column_letter(column)].width = width

    if not accounts:
        sheet.cell(row=_HEADER_ROW + 1, column=1, value=labels["empty"])
    else:
        for account in accounts:
            _write_account(sheet, account, labels)
        _write_grand_total(sheet, accounts, labels)

    sheet.freeze_panes = f"A{_HEADER_ROW + 1}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_account(
    sheet: Worksheet, account: AccountLedger, labels: dict[str, str]
) -> None:
    heading = sheet.max_row + 1
    sheet.cell(
        row=heading, column=1, value=f"{account.code}  {account.name}"
    ).font = _BOLD

    for entry in account.entries:
        row = sheet.max_row + 1
        date = sheet.cell(row=row, column=1, value=entry.date)
        date.number_format = DATE_FORMAT
        sheet.cell(row=row, column=2, value=entry.voucher_number or entry.voucher_id)
        sheet.cell(row=row, column=3, value=entry.description)
        sheet.cell(row=row, column=4, value=entry.third_party_name)
        _money_cell(sheet, row, 5, entry.debit)
        _money_cell(sheet, row, 6, entry.credit)
        _money_cell(sheet, row, 7, entry.running_balance)

    total = sheet.max_row + 1
    sheet.cell(row=total, column=4, value=labels["total"]).font = _BOLD
    for column, amount in enumerate(
        (account.debit, account.credit, account.closing_balance), start=5
    ):
        _money_cell(sheet, total, column, amount).font = _BOLD


def _write_grand_total(
    sheet: Worksheet, accounts: list[AccountLedger], labels: dict[str, str]
) -> None:
    row = sheet.max_row + 2
    sheet.cell(row=row, column=4, value=labels["grand_total"]).font = _BOLD

    for column, amounts in enumerate(
        (
            (a.debit for a in accounts),
            (a.credit for a in accounts),
            (a.closing_balance for a in accounts),
        ),
        start=5,
    ):
        _money_cell(sheet, row, column, sum(amounts, Decimal(0))).font = _BOLD


def _money_cell(sheet: Worksheet, row: int, column: int, value: Decimal) -> Cell:
    cell = sheet.cell(row=row, column=column, value=value)
    cell.number_format = MONEY_FORMAT
    return cell
