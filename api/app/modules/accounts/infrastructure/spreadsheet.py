"""openpyxl adapter for `SpreadsheetReader`.

Keeping the library behind this port is what lets the import use case be tested
with a plain list of tuples, without building a workbook.
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from typing import IO, Final

from openpyxl import load_workbook

from app.shared.errors import DomainError

_HEADER_ROWS: Final = 1
_COLUMNS: Final = 4


class SpreadsheetError(DomainError):
    """The file cannot be read as a chart of accounts."""


@dataclass(frozen=True, slots=True)
class Row:
    number: int
    values: tuple[object, object, object, object]


class OpenpyxlSpreadsheetReader:
    def rows(self, file: IO[bytes]) -> Iterator[Row]:
        try:
            workbook = load_workbook(file, read_only=True, data_only=True)
        except Exception as exc:
            raise SpreadsheetError("The file could not be read as an .xlsx") from exc

        try:
            worksheet = workbook.worksheets[0]
            for number, values in enumerate(
                worksheet.iter_rows(min_row=_HEADER_ROWS + 1, values_only=True),
                start=_HEADER_ROWS + 1,
            ):
                # Padded so short rows still unpack into the four columns.
                padded = (*values, None, None, None, None)[:_COLUMNS]
                yield Row(number=number, values=padded)  # type: ignore[arg-type]
        finally:
            workbook.close()
