"""Import of the chart of accounts from a spreadsheet.

The sheet holds one row per account with the columns `Codigo, Nombre, Tipo,
Naturaleza`. `Tipo` is redundant — the level is derived from the code — so it is
only used to flag inconsistencies, never as the source of truth.

The whole file is one transaction: a chart of accounts committed halfway is
worse than one not imported at all.
"""

from __future__ import annotations

import enum
from dataclasses import dataclass
from typing import IO, Final

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.accounts.cache import AccountTreeCache
from app.modules.accounts.models import Account
from app.modules.accounts.puc import (
    AccountLevel,
    InvalidAccountCode,
    Nature,
    depth_of,
    level_of,
    parent_code_of,
    validate_code,
)
from app.modules.accounts.schemas import ImportResult, RowError
from app.shared.errors import DomainError

_HEADER_ROWS: Final = 1
_COLUMNS: Final = 4
#: Bounds the parameter count of the existence lookup.
_CHUNK: Final = 500


class SpreadsheetError(DomainError):
    """The file cannot be read as a chart of accounts."""


class ExistingAccounts(enum.StrEnum):
    """What to do with accounts that are already stored."""

    SKIP = "skip"
    UPDATE = "update"


@dataclass(frozen=True, slots=True)
class ParsedRow:
    number: int
    code: str
    name: str
    nature: Nature


class AccountImporter:
    def __init__(self, session: AsyncSession, cache: AccountTreeCache) -> None:
        self._session = session
        self._cache = cache

    async def run(
        self,
        file: IO[bytes],
        *,
        on_existing: ExistingAccounts = ExistingAccounts.SKIP,
    ) -> ImportResult:
        rows, errors = _parse(file)

        # Shallowest first, so a parent always lands before its children
        # whatever order the file happens to use.
        rows.sort(key=lambda row: (depth_of(row.code), row.code))

        known = await self._known_codes([row.code for row in rows])
        incoming = {row.code for row in rows}
        stored = (
            await self._load(known) if on_existing is ExistingAccounts.UPDATE else {}
        )

        created = updated = skipped = 0

        for row in rows:
            parent_code = parent_code_of(row.code)
            if parent_code is not None and parent_code not in incoming | known:
                errors.append(
                    RowError(
                        row=row.number,
                        code=row.code,
                        message=f"Missing parent account {parent_code}",
                    )
                )
                continue

            if row.code in known:
                if on_existing is ExistingAccounts.SKIP:
                    skipped += 1
                    continue

                account = stored[row.code]
                account.name = row.name
                account.nature = row.nature
                # Re-importing an account restores it: the file is the authority
                # on which accounts the chart contains.
                account.restore()
                updated += 1
                continue

            self._session.add(
                Account.open(code=row.code, name=row.name, nature=row.nature)
            )
            created += 1

        # One commit for the whole file.
        await self._session.commit()
        if created or updated:
            await self._cache.clear()

        return ImportResult(
            created=created, updated=updated, skipped=skipped, errors=errors
        )

    async def _known_codes(self, codes: list[str]) -> set[str]:
        """Which codes are already stored, chunked to bound the IN() size.

        Soft-deleted rows count: they still occupy the primary key.
        """
        found: set[str] = set()

        for start in range(0, len(codes), _CHUNK):
            result = await self._session.execute(
                select(Account.code).where(
                    Account.code.in_(codes[start : start + _CHUNK])
                )
            )
            found.update(result.scalars().all())

        return found

    async def _load(self, codes: set[str]) -> dict[str, Account]:
        loaded: dict[str, Account] = {}
        ordered = list(codes)

        for start in range(0, len(ordered), _CHUNK):
            result = await self._session.execute(
                select(Account).where(Account.code.in_(ordered[start : start + _CHUNK]))
            )
            loaded.update({row.code: row for row in result.scalars().all()})

        return loaded


def _parse(file: IO[bytes]) -> tuple[list[ParsedRow], list[RowError]]:
    rows: list[ParsedRow] = []
    errors: list[RowError] = []
    seen: set[str] = set()

    for number, values in _iter_rows(file):
        raw_code, raw_name, raw_type, raw_nature = values

        if raw_code is None or str(raw_code).strip() == "":
            continue

        code = str(raw_code).strip()
        try:
            code = validate_code(code)
        except InvalidAccountCode as exc:
            errors.append(RowError(row=number, code=code, message=str(exc)))
            continue

        if code in seen:
            errors.append(
                RowError(row=number, code=code, message="Duplicate code in the file")
            )
            continue

        name = str(raw_name).strip() if raw_name is not None else ""
        if not name:
            errors.append(RowError(row=number, code=code, message="Missing name"))
            continue

        nature = _parse_nature(raw_nature)
        if nature is None:
            errors.append(
                RowError(
                    row=number,
                    code=code,
                    message=f"Unrecognized nature: {raw_nature!r}",
                )
            )
            continue

        if (mismatch := _level_mismatch(code, raw_type)) is not None:
            errors.append(RowError(row=number, code=code, message=mismatch))
            continue

        seen.add(code)
        rows.append(ParsedRow(number=number, code=code, name=name, nature=nature))

    return rows, errors


def _iter_rows(file: IO[bytes]) -> list[tuple[int, tuple[object, ...]]]:
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        raise SpreadsheetError("The file could not be read as an .xlsx") from exc

    try:
        worksheet = workbook.worksheets[0]
        return [
            # Padded so short rows still unpack into the four columns.
            (number, (*values, None, None, None, None)[:_COLUMNS])
            for number, values in enumerate(
                worksheet.iter_rows(min_row=_HEADER_ROWS + 1, values_only=True),
                start=_HEADER_ROWS + 1,
            )
        ]
    finally:
        workbook.close()


def _parse_nature(raw: object) -> Nature | None:
    if raw is None:
        return None

    # The file ships "Debito" and "Crédito"; any accent or casing is accepted so
    # hand-edited sheets keep working.
    normalized = str(raw).strip().lower().replace("é", "e").replace("í", "i")

    if normalized.startswith("deb"):
        return Nature.DEBIT
    if normalized.startswith("cred"):
        return Nature.CREDIT
    return None


def _level_mismatch(code: str, raw_type: object) -> str | None:
    """Flag a `Tipo` column that contradicts the code, which is authoritative."""
    if raw_type is None or str(raw_type).strip() == "":
        return None

    declared = str(raw_type).strip().lower()
    actual = level_of(code)
    if declared == actual.value.lower():
        return None

    if declared not in {level.value.lower() for level in AccountLevel}:
        return f"Unrecognized type: {raw_type!r}"

    return (
        f"Declared type ({raw_type}) does not match the one implied by a "
        f"{len(code)}-digit code ({actual.value})"
    )
