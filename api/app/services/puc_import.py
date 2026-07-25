"""Import of the chart of accounts from a spreadsheet.

The sheet holds one row per account with the columns `Codigo, Nombre, Tipo,
Naturaleza`. `Tipo` is redundant — the level is derived from the code — so it is
only used to flag inconsistencies, never as the source of truth.
"""

from __future__ import annotations

import enum
from collections.abc import Iterator
from dataclasses import dataclass
from typing import IO, Final

from openpyxl import load_workbook

from app.domain.puc import (
    AccountLevel,
    InvalidAccountCode,
    Nature,
    depth_of,
    level_of,
    parent_code_of,
    validate_code,
)
from app.models.account import Account
from app.repositories.account import AccountRepository
from app.schemas.account import ImportResult, RowError

_HEADER_ROWS: Final = 1
_EXPECTED_COLUMNS: Final = 4


class ExistingAccounts(enum.StrEnum):
    """What to do with accounts that are already stored."""

    SKIP = "skip"
    UPDATE = "update"


@dataclass(frozen=True, slots=True)
class ParsedRow:
    """A validated row, ready to become an account."""

    row_number: int
    code: str
    name: str
    nature: Nature


class SpreadsheetError(ValueError):
    """The file cannot be read as a chart of accounts."""


class PucImporter:
    def __init__(self, repository: AccountRepository) -> None:
        self._repository = repository

    async def import_from_file(
        self,
        file: IO[bytes],
        *,
        on_existing: ExistingAccounts = ExistingAccounts.SKIP,
    ) -> ImportResult:
        rows, errors = self._parse(file)

        # Inserting shallowest first guarantees a parent exists before its
        # children, whatever order the file happens to use.
        rows.sort(key=lambda row: (depth_of(row.code), row.code))

        # Soft-deleted rows still hold the primary key, so they count as known.
        known = await self._repository.existing_codes(
            (row.code for row in rows), include_deleted=True
        )
        incoming = {row.code for row in rows}

        created = updated = skipped = 0
        to_insert: list[Account] = []

        for row in rows:
            parent_code = parent_code_of(row.code)
            if parent_code is not None and parent_code not in known | incoming:
                errors.append(
                    RowError(
                        row=row.row_number,
                        code=row.code,
                        message=f"Missing parent account {parent_code}",
                    )
                )
                continue

            if row.code in known:
                if on_existing is ExistingAccounts.SKIP:
                    skipped += 1
                    continue

                account = await self._repository.get(row.code, include_deleted=True)
                if account is not None:
                    account.name = row.name
                    account.nature = row.nature
                    # Re-importing an account restores it: the file is the
                    # authority on which accounts the chart contains.
                    account.deleted_at = None
                    updated += 1
                continue

            to_insert.append(
                Account(
                    code=row.code,
                    name=row.name,
                    nature=row.nature,
                    level=level_of(row.code),
                    parent_code=parent_code,
                )
            )
            created += 1

        self._repository.add_all(to_insert)
        await self._repository.commit()

        return ImportResult(
            created=created, updated=updated, skipped=skipped, errors=errors
        )

    def _parse(self, file: IO[bytes]) -> tuple[list[ParsedRow], list[RowError]]:
        rows: list[ParsedRow] = []
        errors: list[RowError] = []
        seen: set[str] = set()

        for number, values in self._iter_rows(file):
            raw_code, raw_name, raw_type, raw_nature = values

            if raw_code is None or str(raw_code).strip() == "":
                continue

            code = str(raw_code).strip()
            try:
                code = validate_code(code)
            except InvalidAccountCode as exc:
                errors.append(RowError(row=number, code=code, message=str(exc)))
                continue

            if code in seen:
                errors.append(
                    RowError(
                        row=number, code=code, message="Duplicate code in the file"
                    )
                )
                continue

            name = str(raw_name).strip() if raw_name is not None else ""
            if not name:
                errors.append(RowError(row=number, code=code, message="Missing name"))
                continue

            nature = self._parse_nature(raw_nature)
            if nature is None:
                errors.append(
                    RowError(
                        row=number,
                        code=code,
                        message=f"Unrecognized nature: {raw_nature!r}",
                    )
                )
                continue

            if (mismatch := self._level_mismatch(code, raw_type)) is not None:
                errors.append(RowError(row=number, code=code, message=mismatch))
                continue

            seen.add(code)
            rows.append(
                ParsedRow(row_number=number, code=code, name=name, nature=nature)
            )

        return rows, errors

    def _iter_rows(self, file: IO[bytes]) -> Iterator[tuple[int, tuple[object, ...]]]:
        try:
            workbook = load_workbook(file, read_only=True, data_only=True)
        except Exception as exc:
            raise SpreadsheetError("The file could not be read as an .xlsx") from exc

        try:
            worksheet = workbook.worksheets[0]
            for number, values in enumerate(
                worksheet.iter_rows(min_row=_HEADER_ROWS + 1, values_only=True),
                start=_HEADER_ROWS + 1,
            ):
                # Pad so short rows still unpack into the four expected columns.
                padded = (*values, None, None, None, None)[:_EXPECTED_COLUMNS]
                yield number, padded
        finally:
            workbook.close()

    @staticmethod
    def _parse_nature(raw: object) -> Nature | None:
        if raw is None:
            return None

        # The file ships "Debito" and "Crédito"; any accent or casing is accepted
        # so hand-edited sheets keep working.
        normalized = str(raw).strip().lower().replace("é", "e").replace("í", "i")

        if normalized.startswith("deb"):
            return Nature.DEBIT
        if normalized.startswith("cred"):
            return Nature.CREDIT
        return None

    @staticmethod
    def _level_mismatch(code: str, raw_type: object) -> str | None:
        """Flag a `Tipo` column that contradicts the code, which is authoritative."""
        if raw_type is None or str(raw_type).strip() == "":
            return None

        declared = str(raw_type).strip().lower()
        actual = level_of(code)
        if declared == actual.value.lower():
            return None

        known = {level.value.lower() for level in AccountLevel}
        if declared not in known:
            return f"Unrecognized type: {raw_type!r}"

        return (
            f"Declared type ({raw_type}) does not match the one implied by a "
            f"{len(code)}-digit code ({actual.value})"
        )
