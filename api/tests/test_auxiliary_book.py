"""The auxiliary book: what the spreadsheet actually contains."""

from io import BytesIO
from typing import Any

from httpx import AsyncClient
from openpyxl import load_workbook

from tests.test_ledger import sale, seed_chart_without_third_party
from tests.test_vouchers import a_posted_voucher

BASE = "/api/v1/ledger/export"

HEADER_ROW = 4


async def test_the_book_is_a_workbook_that_adds_up(auth_client: AsyncClient) -> None:
    await seed_chart_without_third_party(auth_client)
    await a_posted_voucher(auth_client, **sale())
    await a_posted_voucher(auth_client, **sale(date="2026-07-11"))

    response = await auth_client.get(BASE)

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]

    sheet = load_workbook(BytesIO(response.content)).active
    assert sheet is not None
    rows: list[tuple[Any, ...]] = [
        tuple(row) for row in sheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True)
    ]

    # Seven columns: date, voucher, description, third party, debit, credit,
    # balance. The account is not one of them — it heads its own block, so its
    # movements are the rows between that heading and the account's totals.
    heading = next(
        index
        for index, row in enumerate(rows)
        if str(row[0] or "").startswith("110505")
    )
    cash: list[tuple[Any, ...]] = []
    for row in rows[heading + 1 :]:
        if row[3] == "Totales cuenta":
            break
        cash.append(row)

    assert [row[6] for row in cash] == [150000.0, 300000.0]
    assert all(isinstance(row[4], int | float) for row in cash)

    grand = next(row for row in rows if row[3] == "Total general")
    assert (grand[4], grand[5], grand[6]) == (300000.0, 300000.0, 0)
